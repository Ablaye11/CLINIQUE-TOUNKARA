from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from django.db.models import Sum, F, Q, Count
from django.utils import timezone
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from datetime import date, timedelta
import json
from decimal import Decimal, InvalidOperation
import csv
import io

from .models import Fournisseur, Medicament, MouvementStock, Vente, LigneVente

# ============================================================
# AUTHENTICATION VIEWS
# ============================================================

def get_user_role(user):
    """Returns the role of a Django user based on profile data."""
    if user.is_superuser:
        return 'superadmin'
    # Role stored in last_name field (lightweight approach, no extra model needed)
    if user.last_name == 'caissier':
        return 'caissier'
    return 'admin'

def role_required(*roles):
    """Decorator that allows only users with specified roles."""
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('/login/')
            user_role = get_user_role(request.user)
            if user_role not in roles:
                # Caissier trying to access admin pages
                return redirect('gestion_ventes')
            return view_func(request, *args, **kwargs)
        wrapper.__name__ = view_func.__name__
        return wrapper
    return decorator

def api_role_required(*roles):
    """Decorator for API views that returns JSON error instead of redirecting."""
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return JsonResponse({'error': 'Non authentifié'}, status=401)
            user_role = get_user_role(request.user)
            if user_role not in roles:
                return JsonResponse({'error': 'Accès refusé'}, status=403)
            return view_func(request, *args, **kwargs)
        wrapper.__name__ = view_func.__name__
        return wrapper
    return decorator

def login_view(request):
    if request.user.is_authenticated:
        role = get_user_role(request.user)
        if role == 'caissier':
            return redirect('gestion_ventes')
        return redirect('dashboard')
    error = None
    username = ''
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            if user.is_active:
                login(request, user)
                role = get_user_role(user)
                if role == 'caissier':
                    return redirect('gestion_ventes')
                return redirect('dashboard')
            else:
                error = 'Ce compte est désactivé.'
        else:
            error = 'Identifiant ou mot de passe incorrect.'
    return render(request, 'pharmacie/login.html', {'error': error, 'username': username})

def logout_view(request):
    logout(request)
    return redirect('/login/')

# ============================================================
# PAGE VIEWS (with login & role protection)
# ============================================================

@login_required(login_url='/login/')
@role_required('superadmin', 'admin', 'caissier')
def dashboard(request):
    return render(request, 'pharmacie/dashboard.html')

@login_required(login_url='/login/')
@role_required('superadmin', 'admin')
def liste_medicaments(request):
    fournisseurs = Fournisseur.objects.all()
    categories = Medicament.objects.values_list('categorie', flat=True).distinct()
    return render(request, 'pharmacie/medicaments.html', {
        'fournisseurs': fournisseurs,
        'categories': categories
    })

@login_required(login_url='/login/')
@role_required('superadmin', 'admin')
def gestion_stocks(request):
    medicaments = Medicament.objects.all()
    return render(request, 'pharmacie/stocks.html', {'medicaments': medicaments})

@login_required(login_url='/login/')
@role_required('superadmin', 'admin')
def alertes_page(request):
    return render(request, 'pharmacie/alertes.html')

@login_required(login_url='/login/')
@role_required('superadmin', 'admin', 'caissier')
def gestion_ventes(request):
    return render(request, 'pharmacie/ventes.html')

@login_required(login_url='/login/')
@role_required('superadmin', 'admin')
def gestion_fournisseurs(request):
    return render(request, 'pharmacie/fournisseurs.html')

@login_required(login_url='/login/')
@role_required('superadmin', 'admin')
def rapports_page(request):
    return render(request, 'pharmacie/rapports.html')

@login_required(login_url='/login/')
@role_required('superadmin', 'admin')
def inventaire_page(request):
    """Page d'Inventaire et Bilan de la pharmacie."""
    return render(request, 'pharmacie/inventaire.html')

@login_required(login_url='/login/')
@role_required('superadmin')
def utilisateurs_page(request):
    return render(request, 'pharmacie/utilisateurs.html')


# --- API ENDPOINTS (JSON pour le frontend dynamique) ---

@api_role_required('superadmin', 'admin', 'caissier')
def api_dashboard_stats(request):
    aujourdhui = date.today()
    demain = aujourdhui + timedelta(days=1)
    trente_jours = aujourdhui + timedelta(days=30)
    
    # 1. Valeur totale du stock (prix d'achat & prix de vente)
    medicaments = Medicament.objects.all()
    valeur_achat = sum(m.stock * m.prix_achat for m in medicaments)
    valeur_vente = sum(m.stock * m.prix_vente for m in medicaments)
    
    # 2. Nombre total de médicaments différents
    nb_medicaments = medicaments.count()
    
    # 3. Produits en rupture
    ruptures = medicaments.filter(stock__lte=0).count()
    
    # 4. Produits proches de l'expiration (< 30 jours ou déjà expirés)
    expirations_proches = medicaments.filter(date_expiration__lte=trente_jours).count()
    
    # 5. Ventes du jour (hors ventes annulées)
    ventes_du_jour = Vente.objects.filter(
        date_vente__range=(aujourdhui, demain),
        est_annulee=False
    ).aggregate(total_ventes=Sum('total'))['total_ventes'] or 0
    
    # 6. Achats du jour (estimé via les entrées de stock du jour)
    mouvements_entrees = MouvementStock.objects.filter(type_mouvement='ENTREE', date_mouvement__date=aujourdhui)
    achats_du_jour = sum(mouv.quantite * mouv.medicament.prix_achat for mouv in mouvements_entrees)
    
    # 7. Données pour graphiques (ventes des 7 derniers jours)
    graph_jours = []
    graph_ventes = []
    graph_benefices = []
    
    for i in range(6, -1, -1):
        jour_cible = aujourdhui - timedelta(days=i)
        jour_nom = jour_cible.strftime('%a %d')
        ventes_jour = Vente.objects.filter(date_vente__date=jour_cible, est_annulee=False)
        total_jour = ventes_jour.aggregate(total=Sum('total'))['total'] or 0
        
        # Calculer le bénéfice sur les ventes du jour
        lignes = LigneVente.objects.filter(vente__in=ventes_jour)
        benefice_jour = 0
        for lig in lignes:
            benefice_jour += (lig.prix_unitaire - lig.medicament.prix_achat) * lig.quantite
            
        graph_jours.append(jour_nom)
        graph_ventes.append(float(total_jour))
        graph_benefices.append(float(benefice_jour))
        
    return JsonResponse({
        'valeur_stock_achat': float(valeur_achat),
        'valeur_stock_vente': float(valeur_vente),
        'nb_medicaments': nb_medicaments,
        'ruptures': ruptures,
        'expirations_proches': expirations_proches,
        'ventes_du_jour': float(ventes_du_jour),
        'achats_du_jour': float(achats_du_jour),
        'graph': {
            'labels': graph_jours,
            'ventes': graph_ventes,
            'benefices': graph_benefices
        }
    })

@api_role_required('superadmin', 'admin', 'caissier')
def api_medicaments_liste(request):
    query = request.GET.get('q', '')
    cat = request.GET.get('categorie', '')
    statut_filter = request.GET.get('statut', '')
    exp_filter = request.GET.get('exp', '') # 'proche' (30j) ou 'expired'

    aujourdhui = date.today()
    meds = Medicament.objects.all()

    if query:
        meds = meds.filter(Q(nom__icontains=query) | Q(code__icontains=query) | Q(categorie__icontains=query))
    if cat:
        meds = meds.filter(categorie=cat)
        
    # Filtrer par statut
    data = []
    for m in meds:
        statut = m.statut
        if statut_filter and statut.lower() != statut_filter.lower():
            continue
            
        # Check expiration filter
        if exp_filter == 'expired' and m.date_expiration > aujourdhui:
            continue
        elif exp_filter == 'proche':
            trente_jours = aujourdhui + timedelta(days=30)
            if m.date_expiration <= aujourdhui or m.date_expiration > trente_jours:
                continue
        elif exp_filter == '60':
            soixante_jours = aujourdhui + timedelta(days=60)
            if m.date_expiration <= aujourdhui or m.date_expiration > soixante_jours:
                continue
        elif exp_filter == '90':
            quatre_vingt_dix_jours = aujourdhui + timedelta(days=90)
            if m.date_expiration <= aujourdhui or m.date_expiration > quatre_vingt_dix_jours:
                continue

        data.append({
            'id': m.id,
            'code': m.code,
            'nom': m.nom,
            'categorie': m.categorie,
            'stock': m.stock,
            'stock_minimum': m.stock_minimum,
            'prix_achat': float(m.prix_achat),
            'prix_vente': float(m.prix_vente),
            'date_expiration': m.date_expiration.isoformat(),
            'jours_avant_expiration': m.jours_avant_expiration,
            'fournisseur_id': m.fournisseur.id if m.fournisseur else '',
            'fournisseur_nom': m.fournisseur.nom if m.fournisseur else 'Aucun',
            'statut': statut
        })

    sort_by = request.GET.get('sort', 'nom')
    reverse = request.GET.get('reverse', 'false') == 'true'
    
    if sort_by in ['nom', 'stock', 'prix_vente', 'date_expiration']:
        data.sort(key=lambda x: x[sort_by], reverse=reverse)
        
    return JsonResponse({'medicaments': data})

@csrf_exempt
@api_role_required('superadmin', 'admin')
def api_medicament_save(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            med_id = data.get('id')
            
            prix_achat = Decimal(str(data.get('prix_achat', 0)))
            prix_vente = Decimal(str(data.get('prix_vente', 0)))
            stock = int(data.get('stock', 0))
            stock_min = int(data.get('stock_minimum', 5))
            
            fournisseur_id = data.get('fournisseur')
            fournisseur = None
            if fournisseur_id:
                fournisseur = Fournisseur.objects.filter(id=fournisseur_id).first()

            if med_id:
                med = get_object_or_404(Medicament, id=med_id)
                old_stock = med.stock
                med.code = data.get('code')
                med.nom = data.get('nom')
                med.categorie = data.get('categorie')
                med.prix_achat = prix_achat
                med.prix_vente = prix_vente
                med.stock_minimum = stock_min
                med.date_expiration = data.get('date_expiration')
                med.fournisseur = fournisseur
                
                if old_stock != stock:
                    med.stock = stock
                    med.save()
                    MouvementStock.objects.create(
                        medicament=med,
                        type_mouvement='AJUSTEMENT',
                        quantite=stock,
                        motif="Ajustement manuel lors de la modification rapide"
                    )
                else:
                    med.save()
            else:
                med = Medicament.objects.create(
                    code=data.get('code'),
                    nom=data.get('nom'),
                    categorie=data.get('categorie'),
                    stock=stock,
                    stock_minimum=stock_min,
                    prix_achat=prix_achat,
                    prix_vente=prix_vente,
                    date_expiration=data.get('date_expiration'),
                    fournisseur=fournisseur
                )
                if stock > 0:
                    MouvementStock.objects.create(
                        medicament=med,
                        type_mouvement='ENTREE',
                        quantite=stock,
                        motif="Entrée initiale de stock"
                    )

            return JsonResponse({'success': True, 'id': med.id})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)

@csrf_exempt
@api_role_required('superadmin', 'admin')
def api_medicament_delete(request, med_id):
    if request.method == 'POST':
        med = get_object_or_404(Medicament, id=med_id)
        med.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)


# --- API STOCKS ---

@csrf_exempt
@api_role_required('superadmin', 'admin')
def api_stock_mouvement(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            med_id = data.get('medicament')
            type_mouv = data.get('type_mouvement')
            qty = int(data.get('quantite', 0))
            motif = data.get('motif', '')
            numero_lot = data.get('numero_lot', '')  # ✅ NOUVEAU — numéro de lot

            med = get_object_or_404(Medicament, id=med_id)
            
            if type_mouv == 'SORTIE' and med.stock < qty:
                return JsonResponse({'success': False, 'error': 'Stock insuffisant'}, status=400)

            MouvementStock.objects.create(
                medicament=med,
                type_mouvement=type_mouv,
                quantite=qty,
                motif=motif,
                numero_lot=numero_lot or None
            )
            # ✅ FIX CRITIQUE — Rafraîchir depuis la base après que MouvementStock.save() a modifié le stock
            med.refresh_from_db()
            return JsonResponse({'success': True, 'new_stock': med.stock})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'success': False}, status=405)

@api_role_required('superadmin', 'admin')
def api_stock_historique(request):
    mouvements = MouvementStock.objects.all().select_related('medicament')
    data = [{
        'id': m.id,
        'medicament_nom': m.medicament.nom,
        'medicament_code': m.medicament.code,
        'type_mouvement': m.type_mouvement,
        'quantite': m.quantite,
        'date_mouvement': m.date_mouvement.strftime('%d/%m/%Y %H:%M'),
        'motif': m.motif or '-',
        'numero_lot': m.numero_lot or '-',  # ✅ NOUVEAU
    } for m in mouvements]
    return JsonResponse({'historique': data})


# --- API FOURNISSEURS ---

@api_role_required('superadmin', 'admin')
def api_fournisseurs_liste(request):
    if request.method == 'GET':
        fournisseurs = Fournisseur.objects.all().annotate(
            nb_produits=Count('medicament')
        )
        data = [{
            'id': f.id,
            'nom': f.nom,
            'telephone': f.telephone or '-',
            'email': f.email or '-',
            'adresse': f.adresse or '-',
            'nb_produits': f.nb_produits
        } for f in fournisseurs]
        return JsonResponse({'fournisseurs': data})
        
@csrf_exempt
@api_role_required('superadmin', 'admin')
def api_fournisseur_save(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            f_id = data.get('id')
            if f_id:
                fourn = get_object_or_404(Fournisseur, id=f_id)
                fourn.nom = data.get('nom')
                fourn.telephone = data.get('telephone')
                fourn.email = data.get('email')
                fourn.adresse = data.get('adresse')
                fourn.save()
            else:
                fourn = Fournisseur.objects.create(
                    nom=data.get('nom'),
                    telephone=data.get('telephone'),
                    email=data.get('email'),
                    adresse=data.get('adresse')
                )
            return JsonResponse({'success': True, 'id': fourn.id})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
            
@csrf_exempt
@api_role_required('superadmin', 'admin')
def api_fournisseur_delete(request, f_id):
    if request.method == 'POST':
        fourn = get_object_or_404(Fournisseur, id=f_id)
        fourn.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)


# --- API VENTES & POS ---

@csrf_exempt
@transaction.atomic
@api_role_required('superadmin', 'admin', 'caissier')
def api_vente_creer(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            lignes_data = data.get('lignes', [])
            mode_paiement = data.get('mode_paiement', 'ESPECES')
            client_nom = data.get('client_nom', 'Passant')
            client_adresse = data.get('client_adresse', '')
            client_age_raw = data.get('client_age')
            client_telephone = data.get('client_telephone', '')
            
            client_age = None
            if client_age_raw is not None and str(client_age_raw).strip() != '':
                client_age = str(client_age_raw).strip()
            
            if not lignes_data:
                return JsonResponse({'success': False, 'error': 'Le panier est vide'}, status=400)

            # ✅ Vérification des stocks avec SELECT FOR UPDATE (lock pour éviter race conditions)
            for items in lignes_data:
                med = Medicament.objects.select_for_update().get(id=items['medicament_id'])
                qty = int(items['quantite'])
                if med.stock < qty:
                    return JsonResponse({
                        'success': False, 
                        'error': f'Stock insuffisant pour {med.nom} (Disponible: {med.stock})'
                    }, status=400)

            # Créer le numéro de facture unique
            timestamp = timezone.now().strftime('%Y%m%d%H%M%S')
            last_id = Vente.objects.all().order_by('-id').first()
            seq = (last_id.id + 1) if last_id else 1
            num_facture = f"FAC-{timestamp}-{seq}"
            
            # Créer la vente
            vente = Vente.objects.create(
                numero_facture=num_facture,
                client_nom=client_nom,
                client_adresse=client_adresse,
                client_age=client_age,
                client_telephone=client_telephone,
                mode_paiement=mode_paiement,
                total=0
            )
            
            total_vente = Decimal('0.0')
            for items in lignes_data:
                med = Medicament.objects.get(id=items['medicament_id'])
                qty = int(items['quantite'])
                prix = Decimal(str(items['prix_vente']))
                # ✅ NOUVEAU — Remise optionnelle par ligne
                remise = Decimal(str(items.get('remise', 0)))
                
                ligne = LigneVente.objects.create(
                    vente=vente,
                    medicament=med,
                    quantite=qty,
                    prix_unitaire=prix,
                    remise=remise,
                    # total est calculé automatiquement dans LigneVente.save()
                    total=0
                )
                total_vente += ligne.total
                
            vente.total = total_vente
            vente.save()
            
            return JsonResponse({
                'success': True,
                'numero_facture': num_facture,
                'vente_id': vente.id,
                'total': float(total_vente)
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
            
    return JsonResponse({'success': False}, status=405)


# ✅ NOUVEAU — Annulation d'une vente avec restitution du stock
@csrf_exempt
@transaction.atomic
@api_role_required('superadmin', 'admin', 'caissier')
def api_vente_annuler(request, vente_id):
    if request.method == 'POST':
        try:
            vente = get_object_or_404(Vente, id=vente_id)
            
            if vente.est_annulee:
                return JsonResponse({'success': False, 'error': 'Cette vente est déjà annulée'}, status=400)
            
            # Restituer le stock pour chaque ligne de vente
            for ligne in vente.lignes.all():
                MouvementStock.objects.create(
                    medicament=ligne.medicament,
                    type_mouvement='ENTREE',
                    quantite=ligne.quantite,
                    motif=f"Annulation Facture N° {vente.numero_facture}"
                )
            
            # Marquer la vente comme annulée (on ne supprime pas pour garder la traçabilité)
            vente.est_annulee = True
            vente.save()
            
            return JsonResponse({'success': True, 'message': f'Vente {vente.numero_facture} annulée. Stock restitué.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'success': False, 'error': 'Méthode invalide'}, status=405)


# ✅ NOUVEAU — Remise à zéro complète des données de test (médicaments, ventes, stocks, fournisseurs)
@csrf_exempt
@transaction.atomic
@api_role_required('superadmin')
def api_reinitialiser_donnees(request):
    """Purge toutes les données de test en conservant les comptes utilisateurs."""
    if request.method == 'POST':
        try:
            LigneVente.objects.all().delete()
            Vente.objects.all().delete()
            MouvementStock.objects.all().delete()
            Medicament.objects.all().delete()
            Fournisseur.objects.all().delete()
            return JsonResponse({'success': True, 'message': 'Toutes les données de test ont été effacées. La base est prête pour les données réelles !'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'success': False, 'error': 'Méthode invalide'}, status=405)



@api_role_required('superadmin', 'admin', 'caissier')
def api_ventes_historique(request):
    # ✅ AMÉLIORATION — Filtre optionnel par période
    periode = request.GET.get('periode', 'tout')
    aujourdhui = date.today()
    
    ventes_qs = Vente.objects.all().prefetch_related('lignes__medicament')
    
    if periode == 'jour':
        ventes_qs = ventes_qs.filter(date_vente__date=aujourdhui)
    elif periode == 'semaine':
        ventes_qs = ventes_qs.filter(date_vente__date__gte=aujourdhui - timedelta(days=7))
    elif periode == 'mois':
        ventes_qs = ventes_qs.filter(date_vente__date__gte=aujourdhui - timedelta(days=30))
    elif periode == 'an':
        ventes_qs = ventes_qs.filter(date_vente__date__gte=aujourdhui - timedelta(days=365))
    
    data = []
    for v in ventes_qs:
        lignes = [{
            'medicament_nom': l.medicament.nom,
            'quantite': l.quantite,
            'prix_unitaire': float(l.prix_unitaire),
            'remise': float(l.remise),
            'total': float(l.total)
        } for l in v.lignes.all()]
        
        data.append({
            'id': v.id,
            'numero_facture': v.numero_facture,
            'client_nom': v.client_nom,
            'client_adresse': v.client_adresse or '-',
            'client_age': v.client_age or '-',
            'client_telephone': v.client_telephone or '-',
            'date_vente': v.date_vente.strftime('%d/%m/%Y %H:%M'),
            'mode_paiement': v.get_mode_paiement_display(),
            'total': float(v.total),
            'est_annulee': v.est_annulee,
            'lignes': lignes
        })
    return JsonResponse({'ventes': data})


@login_required(login_url='/login/')
@role_required('superadmin', 'admin', 'caissier')
def api_facture_pdf(request, vente_id):
    vente = get_object_or_404(Vente, id=vente_id)
    return render(request, 'pharmacie/facture_print.html', {'vente': vente})

@csrf_exempt
@api_role_required('superadmin', 'admin')
def api_peupler_db(request):
    if request.method == 'POST':
        try:
            f_pna, _ = Fournisseur.objects.get_or_create(
                nom="Pharmacie Nationale d'Approvisionnement (PNA)",
                defaults={'telephone': "+221 33 821 00 00", 'email': "contact@pna.sn", 'adresse': "Dakar, Sénégal"}
            )
            f_sanofi, _ = Fournisseur.objects.get_or_create(
                nom="Sanofi Sénégal",
                defaults={'telephone': "+221 33 864 12 00", 'email': "info@sanofi.sn", 'adresse': "Almadies, Dakar"}
            )
            f_valpan, _ = Fournisseur.objects.get_or_create(
                nom="Laboratoire Valpan / Medis",
                defaults={'telephone': "+221 33 832 45 67", 'email': "ventes@valpan.sn", 'adresse': "Hann Maristes, Dakar"}
            )
            f_laborex, _ = Fournisseur.objects.get_or_create(
                nom="Laborex Sénégal",
                defaults={'telephone': "+221 33 859 10 10", 'email': "contact@laborex.sn", 'adresse': "KM 4.5 Bd du Centenaire, Dakar"}
            )

            aujourdhui = date.today()

            produits = [
                # Analgésiques & Anti-pyrétiques
                {'code': '3011234567890', 'nom': 'Doliprane 1000mg (Paracétamol)', 'categorie': 'Analgésique', 'stock': 250, 'stock_minimum': 15, 'prix_achat': 500.0, 'prix_vente': 750.0, 'date_expiration': aujourdhui + timedelta(days=730), 'fournisseur': f_sanofi},
                {'code': '3011234567891', 'nom': 'Paracétamol Mylan 500mg', 'categorie': 'Analgésique', 'stock': 120, 'stock_minimum': 15, 'prix_achat': 300.0, 'prix_vente': 450.0, 'date_expiration': aujourdhui + timedelta(days=365), 'fournisseur': f_pna},
                {'code': '3011234567898', 'nom': 'Efferalgan 500mg Comprimés', 'categorie': 'Analgésique', 'stock': 90, 'stock_minimum': 10, 'prix_achat': 450.0, 'prix_vente': 700.0, 'date_expiration': aujourdhui + timedelta(days=500), 'fournisseur': f_laborex},
                {'code': '3011234567899', 'nom': 'Dafalgan Codeine Comprimés', 'categorie': 'Analgésique', 'stock': 45, 'stock_minimum': 5, 'prix_achat': 1100.0, 'prix_vente': 1650.0, 'date_expiration': aujourdhui + timedelta(days=400), 'fournisseur': f_sanofi},

                # Anti-inflammatoires
                {'code': '3011234567893', 'nom': 'Ibuprofène Biogaran 400mg', 'categorie': 'Anti-inflammatoire', 'stock': 85, 'stock_minimum': 10, 'prix_achat': 400.0, 'prix_vente': 600.0, 'date_expiration': aujourdhui + timedelta(days=180), 'fournisseur': f_pna},
                {'code': '3011234567900', 'nom': 'Kétoprofène Biogaran 100mg', 'categorie': 'Anti-inflammatoire', 'stock': 60, 'stock_minimum': 8, 'prix_achat': 950.0, 'prix_vente': 1400.0, 'date_expiration': aujourdhui + timedelta(days=350), 'fournisseur': f_valpan},
                {'code': '3011234567901', 'nom': 'Diclofénac Sandoz 50mg', 'categorie': 'Anti-inflammatoire', 'stock': 75, 'stock_minimum': 10, 'prix_achat': 550.0, 'prix_vente': 850.0, 'date_expiration': aujourdhui + timedelta(days=420), 'fournisseur': f_laborex},
                {'code': '3011234567902', 'nom': 'Voltarène Emulgel 1%', 'categorie': 'Anti-inflammatoire', 'stock': 30, 'stock_minimum': 5, 'prix_achat': 1800.0, 'prix_vente': 2600.0, 'date_expiration': aujourdhui + timedelta(days=600), 'fournisseur': f_laborex},

                # Antibiotiques
                {'code': '3011234567892', 'nom': 'Amoxicilline Sandoz 500mg', 'categorie': 'Antibiotique', 'stock': 140, 'stock_minimum': 15, 'prix_achat': 1200.0, 'prix_vente': 1800.0, 'date_expiration': aujourdhui + timedelta(days=540), 'fournisseur': f_valpan},
                {'code': '3011234567896', 'nom': 'Augmentin Nourrisson 100mg', 'categorie': 'Antibiotique', 'stock': 40, 'stock_minimum': 5, 'prix_achat': 2200.0, 'prix_vente': 3100.0, 'date_expiration': aujourdhui + timedelta(days=95), 'fournisseur': f_laborex},
                {'code': '3011234567903', 'nom': 'Ciprofloxacine 500mg', 'categorie': 'Antibiotique', 'stock': 65, 'stock_minimum': 8, 'prix_achat': 1400.0, 'prix_vente': 2100.0, 'date_expiration': aujourdhui + timedelta(days=480), 'fournisseur': f_pna},
                {'code': '3011234567904', 'nom': 'Azithromycine Zithromax 500mg', 'categorie': 'Antibiotique', 'stock': 35, 'stock_minimum': 5, 'prix_achat': 2800.0, 'prix_vente': 3900.0, 'date_expiration': aujourdhui + timedelta(days=300), 'fournisseur': f_sanofi},
                {'code': '3011234567905', 'nom': 'Flagyl 500mg (Métronidazole)', 'categorie': 'Antibiotique', 'stock': 110, 'stock_minimum': 10, 'prix_achat': 750.0, 'prix_vente': 1150.0, 'date_expiration': aujourdhui + timedelta(days=620), 'fournisseur': f_sanofi},

                # Antipaludéens
                {'code': '3011234567906', 'nom': 'Coartem 20/120mg (Artéméther)', 'categorie': 'Antipaludéen', 'stock': 180, 'stock_minimum': 20, 'prix_achat': 1500.0, 'prix_vente': 2200.0, 'date_expiration': aujourdhui + timedelta(days=700), 'fournisseur': f_pna},
                {'code': '3011234567907', 'nom': 'Artefan 80/480mg Adulte', 'categorie': 'Antipaludéen', 'stock': 95, 'stock_minimum': 10, 'prix_achat': 1800.0, 'prix_vente': 2700.0, 'date_expiration': aujourdhui + timedelta(days=450), 'fournisseur': f_pna},
                {'code': '3011234567908', 'nom': 'Quinine 300mg Comprimés', 'categorie': 'Antipaludéen', 'stock': 50, 'stock_minimum': 5, 'prix_achat': 900.0, 'prix_vente': 1350.0, 'date_expiration': aujourdhui + timedelta(days=380), 'fournisseur': f_valpan},

                # Gastro-entérologie & Antispasmodiques
                {'code': '3011234567894', 'nom': 'Spasfon Lyoc (Phloroglucinol)', 'categorie': 'Antispasmodique', 'stock': 110, 'stock_minimum': 10, 'prix_achat': 800.0, 'prix_vente': 1200.0, 'date_expiration': aujourdhui + timedelta(days=400), 'fournisseur': f_sanofi},
                {'code': '3011234567895', 'nom': 'Gaviscon Suspension Buvable', 'categorie': 'Anti-acide', 'stock': 45, 'stock_minimum': 5, 'prix_achat': 900.0, 'prix_vente': 1400.0, 'date_expiration': aujourdhui + timedelta(days=450), 'fournisseur': f_laborex},
                {'code': '3011234567909', 'nom': 'Smecta Sachets Poudre', 'categorie': 'Anti-diarrhéique', 'stock': 200, 'stock_minimum': 20, 'prix_achat': 150.0, 'prix_vente': 250.0, 'date_expiration': aujourdhui + timedelta(days=800), 'fournisseur': f_sanofi},
                {'code': '3011234567910', 'nom': 'Maalox Suspension Stomach', 'categorie': 'Anti-acide', 'stock': 55, 'stock_minimum': 5, 'prix_achat': 1100.0, 'prix_vente': 1650.0, 'date_expiration': aujourdhui + timedelta(days=360), 'fournisseur': f_sanofi},
                {'code': '3011234567911', 'nom': 'Oméprazole 20mg Gélules', 'categorie': 'Anti-ulcéreux', 'stock': 130, 'stock_minimum': 15, 'prix_achat': 650.0, 'prix_vente': 1000.0, 'date_expiration': aujourdhui + timedelta(days=520), 'fournisseur': f_valpan},
                {'code': '3011234567912', 'nom': 'Imodium 2mg (Lopéramide)', 'categorie': 'Anti-diarrhéique', 'stock': 70, 'stock_minimum': 10, 'prix_achat': 850.0, 'prix_vente': 1300.0, 'date_expiration': aujourdhui + timedelta(days=490), 'fournisseur': f_laborex},

                # Vitamines & Suppléments
                {'code': '3011234567913', 'nom': 'Vitamine C Laroscorbine 1g', 'categorie': 'Vitamines', 'stock': 160, 'stock_minimum': 15, 'prix_achat': 700.0, 'prix_vente': 1100.0, 'date_expiration': aujourdhui + timedelta(days=650), 'fournisseur': f_sanofi},
                {'code': '3011234567914', 'nom': 'Zinc 20mg Comprimés', 'categorie': 'Vitamines', 'stock': 140, 'stock_minimum': 10, 'prix_achat': 350.0, 'prix_vente': 550.0, 'date_expiration': aujourdhui + timedelta(days=720), 'fournisseur': f_pna},
                {'code': '3011234567915', 'nom': 'Ranferon Fer + Acide Folique', 'categorie': 'Vitamines', 'stock': 85, 'stock_minimum': 10, 'prix_achat': 1250.0, 'prix_vente': 1850.0, 'date_expiration': aujourdhui + timedelta(days=410), 'fournisseur': f_valpan},
                {'code': '3011234567916', 'nom': 'Calcium Sandoz 500mg Efferv.', 'categorie': 'Vitamines', 'stock': 50, 'stock_minimum': 5, 'prix_achat': 1600.0, 'prix_vente': 2400.0, 'date_expiration': aujourdhui + timedelta(days=330), 'fournisseur': f_laborex},

                # Respiratoire & Allergies
                {'code': '3011234567897', 'nom': 'Ventoline HFA Inhalateur', 'categorie': 'Bronchodilatateur', 'stock': 60, 'stock_minimum': 5, 'prix_achat': 1500.0, 'prix_vente': 2300.0, 'date_expiration': aujourdhui + timedelta(days=540), 'fournisseur': f_laborex},
                {'code': '3011234567917', 'nom': 'Cétirizine Biogaran 10mg', 'categorie': 'Antihistaminique', 'stock': 95, 'stock_minimum': 10, 'prix_achat': 500.0, 'prix_vente': 750.0, 'date_expiration': aujourdhui + timedelta(days=580), 'fournisseur': f_valpan},
                {'code': '3011234567918', 'nom': 'Loratadine 10mg Comprimés', 'categorie': 'Antihistaminique', 'stock': 75, 'stock_minimum': 8, 'prix_achat': 450.0, 'prix_vente': 700.0, 'date_expiration': aujourdhui + timedelta(days=610), 'fournisseur': f_pna},
                {'code': '3011234567919', 'nom': 'Humex Rhume Comprimés', 'categorie': 'Respiratoire', 'stock': 65, 'stock_minimum': 10, 'prix_achat': 1300.0, 'prix_vente': 1950.0, 'date_expiration': aujourdhui + timedelta(days=440), 'fournisseur': f_laborex},

                # Antiseptiques & Soins
                {'code': '3011234567920', 'nom': 'Bétadine Dermique 10% (Jaune)', 'categorie': 'Antiseptique', 'stock': 90, 'stock_minimum': 10, 'prix_achat': 1200.0, 'prix_vente': 1750.0, 'date_expiration': aujourdhui + timedelta(days=730), 'fournisseur': f_sanofi},
                {'code': '3011234567921', 'nom': 'Alcool Éthylique 70° 250ml', 'categorie': 'Antiseptique', 'stock': 120, 'stock_minimum': 15, 'prix_achat': 600.0, 'prix_vente': 950.0, 'date_expiration': aujourdhui + timedelta(days=1000), 'fournisseur': f_pna},
                {'code': '3011234567922', 'nom': 'Sérum Physiologique NaCl 0.9%', 'categorie': 'Soins', 'stock': 220, 'stock_minimum': 25, 'prix_achat': 200.0, 'prix_vente': 350.0, 'date_expiration': aujourdhui + timedelta(days=900), 'fournisseur': f_valpan},
                {'code': '3011234567923', 'nom': 'Dacryosérum Collyre 10ml', 'categorie': 'Ophtalmologie', 'stock': 40, 'stock_minimum': 5, 'prix_achat': 850.0, 'prix_vente': 1300.0, 'date_expiration': aujourdhui + timedelta(days=320), 'fournisseur': f_laborex},
            ]

            inserted = 0
            updated = 0
            for prod in produits:
                med, created = Medicament.objects.get_or_create(
                    code=prod['code'],
                    defaults={
                        'nom': prod['nom'], 'categorie': prod['categorie'],
                        'stock': prod['stock'], 'stock_minimum': prod['stock_minimum'],
                        'prix_achat': Decimal(str(prod['prix_achat'])),
                        'prix_vente': Decimal(str(prod['prix_vente'])),
                        'date_expiration': prod['date_expiration'],
                        'fournisseur': prod['fournisseur']
                    }
                )
                if created:
                    inserted += 1
                    if med.stock > 0:
                        MouvementStock.objects.create(
                            medicament=med, type_mouvement='ENTREE',
                            quantite=med.stock, motif="Importation catalogue initiale"
                        )
                else:
                    # Mettre à jour le produit existant
                    med.nom = prod['nom']
                    med.categorie = prod['categorie']
                    med.stock_minimum = prod['stock_minimum']
                    med.prix_achat = Decimal(str(prod['prix_achat']))
                    med.prix_vente = Decimal(str(prod['prix_vente']))
                    med.fournisseur = prod['fournisseur']
                    if med.stock == 0:
                        med.stock = prod['stock']
                        MouvementStock.objects.create(
                            medicament=med, type_mouvement='ENTREE',
                            quantite=med.stock, motif="Réapprovisionnement automatique"
                        )
                    med.save()
                    updated += 1

            return JsonResponse({'success': True, 'inserted': inserted, 'updated': updated, 'total': Medicament.objects.count()})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'success': False}, status=405)


# --- API RAPPORTS DÉTAILLÉS ---

@api_role_required('superadmin', 'admin')
def api_rapports_data(request):
    periode = request.GET.get('periode', 'mois')
    aujourdhui = date.today()
    
    if periode == 'jour':
        start_date = aujourdhui
    elif periode == 'semaine':
        start_date = aujourdhui - timedelta(days=7)
    elif periode == 'an':
        start_date = aujourdhui - timedelta(days=365)
    else:
        start_date = aujourdhui - timedelta(days=30)
        
    # ✅ Exclure les ventes annulées des rapports
    ventes = Vente.objects.filter(date_vente__date__gte=start_date, est_annulee=False)
    
    total_ventes = ventes.aggregate(total=Sum('total'))['total'] or 0
    
    lignes = LigneVente.objects.filter(vente__in=ventes).select_related('medicament')
    total_benefice = 0
    for l in lignes:
        total_benefice += (l.prix_unitaire - l.medicament.prix_achat) * l.quantite
        
    top_meds_query = LigneVente.objects.filter(vente__in=ventes)\
        .values('medicament__nom')\
        .annotate(total_quantite=Sum('quantite'), total_somme=Sum('total'))\
        .order_by('-total_quantite')[:10]
        
    top_meds = [{
        'nom': item['medicament__nom'],
        'quantite': item['total_quantite'],
        'total': float(item['total_somme'])
    } for item in top_meds_query]
    
    cat_query = LigneVente.objects.filter(vente__in=ventes)\
        .values('medicament__categorie')\
        .annotate(total_quantite=Sum('quantite'))\
        .order_by('-total_quantite')
        
    categories_ventes = [{
        'categorie': item['medicament__categorie'],
        'quantite': item['total_quantite']
    } for item in cat_query]
    
    return JsonResponse({
        'periode': periode,
        'total_ventes': float(total_ventes),
        'total_benefice': float(total_benefice),
        'top_medicaments': top_meds,
        'categories_ventes': categories_ventes
    })


# ============================================================
# API INVENTAIRE & BILAN (AUTOMATIQUE)
# ============================================================

@api_role_required('superadmin', 'admin')
def api_inventaire_data(request):
    """Calcul automatique du bilan d'inventaire (valeur stock, entrées, ventes, sorties)."""
    periode = request.GET.get('periode', 'jour') # jour, hier, mois, an, tout
    date_str = request.GET.get('date', '') # YYYY-MM-DD optionnel
    
    aujourdhui = date.today()
    
    if date_str:
        try:
            from datetime import datetime
            start_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            end_date = start_date
        except ValueError:
            start_date = aujourdhui
            end_date = aujourdhui
    elif periode == 'jour':
        start_date = aujourdhui
        end_date = aujourdhui
    elif periode == 'hier':
        start_date = aujourdhui - timedelta(days=1)
        end_date = start_date
    elif periode == 'semaine':
        start_date = aujourdhui - timedelta(days=7)
        end_date = aujourdhui
    elif periode == 'mois':
        start_date = date(aujourdhui.year, aujourdhui.month, 1)
        end_date = aujourdhui
    elif periode == 'an':
        start_date = date(aujourdhui.year, 1, 1)
        end_date = aujourdhui
    else:
        start_date = date(2000, 1, 1)
        end_date = aujourdhui + timedelta(days=365)

    start_datetime = timezone.make_aware(timezone.datetime.combine(start_date, timezone.datetime.min.time()))
    end_datetime = timezone.make_aware(timezone.datetime.combine(end_date, timezone.datetime.max.time()))

    # 1. Tous les médicaments avec stock actuel
    medicaments = Medicament.objects.all().select_related('fournisseur')
    
    valeur_stock_achat_totale = sum(m.stock * m.prix_achat for m in medicaments)
    valeur_stock_vente_totale = sum(m.stock * m.prix_vente for m in medicaments)
    marge_stock_restant_totale = valeur_stock_vente_totale - valeur_stock_achat_totale

    # 2. Ventes de la période (exclure annulées)
    ventes_qs = Vente.objects.filter(date_vente__range=(start_datetime, end_datetime), est_annulee=False)
    somme_vendue_periode = ventes_qs.aggregate(total=Sum('total'))['total'] or 0

    lignes_ventes = LigneVente.objects.filter(vente__in=ventes_qs)
    benefice_ventes_periode = 0
    for lig in lignes_ventes:
        benefice_ventes_periode += (lig.prix_unitaire - lig.medicament.prix_achat) * lig.quantite

    # 3. Mouvements d'Entrée sur la période
    entrees_qs = MouvementStock.objects.filter(type_mouvement='ENTREE', date_mouvement__range=(start_datetime, end_datetime))
    total_entrees_quantite = entrees_qs.aggregate(qte=Sum('quantite'))['qte'] or 0
    valeur_entrees_vente = sum(e.quantite * e.medicament.prix_vente for e in entrees_qs)
    valeur_entrees_achat = sum(e.quantite * e.medicament.prix_achat for e in entrees_qs)

    # 4. Mouvements de Sortie hors-vente (pertes, expirations)
    sorties_hors_vente_qs = MouvementStock.objects.filter(type_mouvement='SORTIE', date_mouvement__range=(start_datetime, end_datetime)).exclude(motif__icontains="Vente Facture")
    total_sorties_quantite = sorties_hors_vente_qs.aggregate(qte=Sum('quantite'))['qte'] or 0
    valeur_sorties_vente = sum(s.quantite * s.medicament.prix_vente for s in sorties_hors_vente_qs)

    # 5. Calcul détaillé par médicament
    inventory_items = []
    for m in medicaments:
        # Entrées sur la période pour ce produit
        m_entrees = MouvementStock.objects.filter(medicament=m, type_mouvement='ENTREE', date_mouvement__range=(start_datetime, end_datetime)).aggregate(q=Sum('quantite'))['q'] or 0
        # Sorties / Ventes sur la période
        m_sorties = MouvementStock.objects.filter(medicament=m, type_mouvement='SORTIE', date_mouvement__range=(start_datetime, end_datetime)).aggregate(q=Sum('quantite'))['q'] or 0

        stock_actuel = m.stock
        valeur_restante_vente = stock_actuel * float(m.prix_vente)
        valeur_restante_achat = stock_actuel * float(m.prix_achat)

        inventory_items.append({
            'id': m.id,
            'code': m.code,
            'nom': m.nom,
            'categorie': m.categorie,
            'prix_achat': float(m.prix_achat),
            'prix_vente': float(m.prix_vente),
            'stock_actuel': stock_actuel,
            'stock_minimum': m.stock_minimum,
            'entrees_periode': m_entrees,
            'sorties_periode': m_sorties,
            'valeur_restante_vente': valeur_restante_vente,
            'valeur_restante_achat': valeur_restante_achat,
            'statut': m.statut,
            'fournisseur_nom': m.fournisseur.nom if m.fournisseur else 'Aucun'
        })

    return JsonResponse({
        'periode': periode,
        'date_debut': start_date.strftime('%d/%m/%Y'),
        'date_fin': end_date.strftime('%d/%m/%Y'),
        'stats': {
            'valeur_stock_vente': float(valeur_stock_vente_totale),
            'valeur_stock_achat': float(valeur_stock_achat_totale),
            'marge_stock_restant': float(marge_stock_restant_totale),
            'somme_vendue': float(somme_vendue_periode),
            'benefice_ventes': float(benefice_ventes_periode),
            'nb_ventes': ventes_qs.count(),
            'total_entrees_qte': total_entrees_quantite,
            'valeur_entrees_achat': float(valeur_entrees_achat),
            'valeur_entrees_vente': float(valeur_entrees_vente),
            'total_sorties_qte': total_sorties_quantite,
            'valeur_sorties_vente': float(valeur_sorties_vente),
            'total_produits': medicaments.count()
        },
        'items': inventory_items
    })


@api_role_required('superadmin', 'admin')
def api_inventaire_export_csv(request):
    """Exporter la fiche d'inventaire au format CSV."""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    today_str = date.today().strftime('%Y-%m-%d')
    response['Content-Disposition'] = f'attachment; filename="inventaire_pharmacie_{today_str}.csv"'
    response.write('\ufeff')

    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Code', 'Médicament', 'Catégorie', 'Stock Restant', 'Prix Achat (FCFA)', 'Prix Vente (FCFA)', 'Valeur Stock Achat', 'Valeur Stock Vente', 'Statut', 'Fournisseur'])

    medicaments = Medicament.objects.all().select_related('fournisseur')
    for m in medicaments:
        val_achat = m.stock * m.prix_achat
        val_vente = m.stock * m.prix_vente
        writer.writerow([
            m.code,
            m.nom,
            m.categorie,
            m.stock,
            m.prix_achat,
            m.prix_vente,
            val_achat,
            val_vente,
            m.statut,
            m.fournisseur.nom if m.fournisseur else 'Aucun'
        ])

    return response


# ============================================================
# IMPORTATION EN MASSE (EXCEL .XLSX / .XLS ET CSV)
# ============================================================

@api_role_required('superadmin', 'admin')
def api_import_template(request):
    """Télécharger un fichier modèle Excel (.xlsx) ou CSV pour l'import de médicaments."""
    fmt = request.GET.get('format', 'excel')
    
    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="modele_import_medicaments.csv"'
        response.write('\ufeff')

        writer = csv.writer(response, delimiter=';')
        # Colonnes : (* = obligatoire)
        writer.writerow(['nom *', 'categorie *', 'prix_achat *', 'prix_vente *', 'stock', 'stock_minimum', 'date_expiration', 'fournisseur_nom', 'code (optionnel)'])
        writer.writerow(['Paracetamol 500mg', 'Analgésique', 500, 750, 100, 10, '2027-12-31', "Pharmacie Nationale d'Approvisionnement (PNA)", ''])
        writer.writerow(['Amoxicilline 500mg', 'Antibiotique', 1200, 1800, 50, 5, '2027-06-30', 'PNA', ''])
        return response

    # Modèle Excel (.xlsx) par défaut
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Modèle Import"

        # Colonnes : obligatoires (rouge) + optionnelles (gris)
        headers = [
            ('nom',            True,  'Nom complet du médicament'),
            ('categorie',      True,  'Catégorie (ex: Antibiotique, Analgésique...)'),
            ('prix_achat',     True,  'Prix d\'achat en FCFA'),
            ('prix_vente',     True,  'Prix de vente en FCFA'),
            ('stock',          False, 'Quantité en stock (0 par défaut)'),
            ('stock_minimum',  False, 'Seuil d\'alerte stock faible (5 par défaut)'),
            ('date_expiration',False, 'Date d\'expiration (AAAA-MM-JJ, optionnel)'),
            ('fournisseur_nom',False, 'Nom exact du fournisseur (optionnel)'),
            ('code',           False, 'Code-barres (auto-généré si vide)'),
        ]

        # Style en-têtes obligatoires (rouge foncé)
        fill_required = PatternFill(start_color="B91C1C", end_color="B91C1C", fill_type="solid")
        # Style en-têtes optionnels (gris bleué)
        fill_optional = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
        font_header   = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        font_optional = Font(name="Arial", size=11, bold=False, color="D1D5DB")
        align_center  = Alignment(horizontal="center", vertical="center", wrap_text=True)

        col_names = [h[0] for h in headers]
        ws.append(col_names)
        ws.row_dimensions[1].height = 28

        for col_num, (col_name, is_required, comment) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill   = fill_required if is_required else fill_optional
            cell.font   = font_header if is_required else font_optional
            cell.alignment = align_center
            # Ajouter un commentaire explicatif
            from openpyxl.comments import Comment as XlComment
            label = '(*) OBLIGATOIRE' if is_required else '(optionnel)'
            cell.comment = XlComment(f"{label}\n{comment}", "Clinique Tounkara")

        # Ligne d'exemple 1
        ws.append([
            'Paracetamol 500mg', 'Analgésique', 500, 750,
            100, 10, '2027-12-31',
            "Pharmacie Nationale d'Approvisionnement (PNA)", ''
        ])
        # Ligne d'exemple 2
        ws.append([
            'Amoxicilline 500mg', 'Antibiotique', 1200, 1800,
            50, 5, '2027-06-30', 'PNA', ''
        ])
        # Ligne d'exemple 3
        ws.append([
            'Doliprane 1000mg', 'Analgésique', 600, 850,
            80, 10, '2028-01-15', '', ''
        ])

        # Style lignes d'exemples
        fill_example = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        font_example = Font(name="Arial", size=10, italic=True, color="9CA3AF")
        for row_idx in range(2, 5):
            for col_idx in range(1, len(headers)+1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.fill = fill_example
                c.font = font_example
                c.alignment = Alignment(horizontal="left", vertical="center")

        # Ajuster largeur colonnes
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 6, 20)

        # Figer la ligne d'en-tête
        ws.freeze_panes = 'A2'

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="modele_import_medicaments.xlsx"'
        wb.save(response)
        return response
    except ImportError:
        # Fallback CSV si openpyxl n'est pas présent
        return redirect('/api/import/template/?format=csv')


@csrf_exempt
@api_role_required('superadmin', 'admin')
def api_import_csv(request):
    """Importer des médicaments depuis un fichier Excel (.xlsx / .xls) ou CSV."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode invalide'}, status=405)

    uploaded_file = request.FILES.get('fichier_csv') or request.FILES.get('fichier_excel') or request.FILES.get('file')
    if not uploaded_file:
        return JsonResponse({'success': False, 'error': 'Aucun fichier reçu'}, status=400)

    filename = uploaded_file.name.lower()
    rows_data = []

    if filename.endswith(('.xlsx', '.xls')):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(uploaded_file, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            
            if not all_rows or len(all_rows) < 2:
                return JsonResponse({'success': False, 'error': 'Le fichier Excel est vide ou ne contient pas de données.'}, status=400)

            headers_raw = [str(cell).strip().lower() if cell is not None else '' for cell in all_rows[0]]
            # Normaliser : supprimer *, (optionnel), espaces
            headers = [
                h.replace(' *', '').replace('*', '')
                 .replace(' (optionnel)', '').replace('(optionnel)', '')
                 .strip()
                for h in headers_raw
            ]
            
            for row in all_rows[1:]:
                if not any(row):
                    continue
                row_dict = {}
                for idx, val in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        row_dict[headers[idx]] = str(val).strip() if val is not None else ''
                rows_data.append(row_dict)
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Erreur de lecture du fichier Excel: {str(e)}'}, status=400)
    else:
        # Fichier CSV
        try:
            raw_content = uploaded_file.read()
            try:
                content = raw_content.decode('utf-8-sig')
            except UnicodeDecodeError:
                content = raw_content.decode('latin-1')
            
            first_line = content.split('\n')[0]
            delimiter = ';' if first_line.count(';') >= first_line.count(',') else ','
            reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
            
            for row in reader:
                # Normaliser les clés : supprimer *, (optionnel), espaces
                row_dict = {
                    k.strip().lower()
                     .replace(' *', '').replace('*', '')
                     .replace(' (optionnel)', '').replace('(optionnel)', '')
                     .strip(): v.strip()
                    for k, v in row.items() if k
                }
                rows_data.append(row_dict)
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Erreur de lecture du fichier CSV: {str(e)}'}, status=400)

    if not rows_data:
        return JsonResponse({'success': False, 'error': 'Aucune ligne valide trouvée dans le fichier.'}, status=400)

    # ── Validation des colonnes obligatoires ──
    required_cols = {'nom', 'prix_achat', 'prix_vente'}
    actual_cols = set(rows_data[0].keys())
    missing = required_cols - actual_cols
    if missing:
        return JsonResponse({
            'success': False,
            'error': f'Colonnes obligatoires manquantes : {", ".join(sorted(missing))}. Colonnes requises : nom, categorie, prix_achat, prix_vente. Téléchargez le modèle pour la structure exacte.'
        }, status=400)

    inserted = 0
    updated = 0
    errors = []

    for i, row in enumerate(rows_data, start=2):
        nom       = row.get('nom', '').strip()
        categorie = row.get('categorie', '').strip() or 'Général'

        if not nom:
            errors.append(f"Ligne {i} ignorée : le nom du médicament est obligatoire.")
            continue

        try:
            stock       = int(float(row.get('stock', 0) or 0))
            stock_min   = int(float(row.get('stock_minimum', 5) or 5))
            prix_achat  = Decimal(str(row.get('prix_achat', 0) or 0).replace(',', '.'))
            prix_vente  = Decimal(str(row.get('prix_vente', 0) or 0).replace(',', '.'))
        except (ValueError, InvalidOperation) as e:
            errors.append(f"Ligne {i} ({nom}) ignorée : valeur numérique invalide → {e}")
            continue

        if prix_achat <= 0 or prix_vente <= 0:
            errors.append(f"Ligne {i} ({nom}) ignorée : prix_achat et prix_vente doivent être supérieurs à 0.")
            continue

        date_exp_str = row.get('date_expiration', '').strip()
        date_exp = None
        
        if date_exp_str and date_exp_str not in ('', 'None', 'nan'):
            # Traiter les dates au format datetime Excel (ex: '2027-12-31 00:00:00')
            if ' ' in date_exp_str:
                date_exp_str = date_exp_str.split(' ')[0]
                
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
                try:
                    from datetime import datetime
                    date_exp = datetime.strptime(date_exp_str, fmt).date()
                    break
                except ValueError:
                    continue
                    
        if date_exp is None:
            # Date par défaut dans 2 ans si non spécifiée ou invalide
            date_exp = date.today() + timedelta(days=730)

        fournisseur = None
        fourn_nom = row.get('fournisseur_nom', '').strip()
        if fourn_nom and fourn_nom not in ('', 'None', 'nan'):
            fournisseur, _ = Fournisseur.objects.get_or_create(nom=fourn_nom)

        # ── Auto-génération du code si absent ──
        code = row.get('code', '').strip()
        if not code or code in ('', 'None', 'nan'):
            import hashlib
            # Code unique basé sur nom + catégorie + prix_vente
            seed = f"{nom.lower().strip()}-{categorie.lower().strip()}-{prix_vente}"
            code = 'AUTO-' + hashlib.md5(seed.encode()).hexdigest()[:10].upper()

        # ── Chercher d'abord par code, puis par nom+catégorie ──
        med, created = Medicament.objects.get_or_create(
            code=code,
            defaults={
                'nom': nom, 'categorie': categorie, 'stock': stock,
                'stock_minimum': stock_min, 'prix_achat': prix_achat,
                'prix_vente': prix_vente, 'date_expiration': date_exp,
                'fournisseur': fournisseur,
            }
        )

        if created:
            inserted += 1
            if stock > 0:
                MouvementStock.objects.create(
                    medicament=med, type_mouvement='ENTREE',
                    quantite=stock, motif='Importation initiale (Excel/CSV)'
                )
        else:
            med.nom = nom
            med.categorie = categorie
            med.stock_minimum = stock_min
            med.prix_achat = prix_achat
            med.prix_vente = prix_vente
            med.date_expiration = date_exp
            if fournisseur:
                med.fournisseur = fournisseur
            old_stock = med.stock
            med.stock = stock
            med.save()
            if old_stock != stock:
                MouvementStock.objects.create(
                    medicament=med, type_mouvement='AJUSTEMENT',
                    quantite=stock, motif='Mise à jour via import (Excel/CSV)'
                )
            updated += 1


    return JsonResponse({
        'success': True,
        'inserted': inserted,
        'updated': updated,
        'errors': errors,
        'total_processed': inserted + updated
    })



# ============================================================
# API GESTION DES UTILISATEURS (Super-Admin only)
# ============================================================

@api_role_required('superadmin')
def api_utilisateurs_liste(request):
    users = User.objects.all().order_by('id')
    data = []
    for u in users:
        role = get_user_role(u)
        display_name = f"{u.first_name} {u.last_name}".strip() if u.last_name not in ('admin', 'caissier') else u.first_name
        if not display_name:
            display_name = u.username
        data.append({
            'id': u.id,
            'nom': display_name,
            'username': u.username,
            'role': role,
            'is_active': u.is_active,
        })
    return JsonResponse({'users': data})

@csrf_exempt
@api_role_required('superadmin')
def api_utilisateur_save(request):
    if request.method != 'POST':
        return JsonResponse({'success': False}, status=405)
    data = json.loads(request.body)
    user_id = data.get('id')
    nom = data.get('nom', '').strip()
    username = data.get('username', '').strip()
    role = data.get('role', 'admin')
    password = data.get('password', '').strip()

    if not nom or not username:
        return JsonResponse({'success': False, 'error': 'Nom et identifiant requis.'})

    # Split nom into first/last for storage
    parts = nom.split(' ', 1)
    first_name = parts[0]
    last_name_display = parts[1] if len(parts) > 1 else ''

    if user_id:
        # Update existing user
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Utilisateur introuvable.'})
        if user.is_superuser:
            return JsonResponse({'success': False, 'error': 'Le super-administrateur ne peut pas être modifié ici.'})
        # Check username uniqueness
        if User.objects.filter(username=username).exclude(id=user_id).exists():
            return JsonResponse({'success': False, 'error': f"L'identifiant '{username}' est déjà utilisé."})
        user.first_name = first_name
        # Store role in last_name field (simple approach)
        user.last_name = role  # 'admin' or 'caissier'
        user.username = username
        user.is_staff = (role == 'admin')
        if password:
            if len(password) < 4:
                return JsonResponse({'success': False, 'error': 'Le mot de passe doit contenir au moins 4 caractères.'})
            user.set_password(password)
        user.save()
        return JsonResponse({'success': True})
    else:
        # Create new user
        if User.objects.filter(username=username).exists():
            return JsonResponse({'success': False, 'error': f"L'identifiant '{username}' est déjà utilisé."})
        if not password or len(password) < 4:
            return JsonResponse({'success': False, 'error': 'Le mot de passe doit contenir au moins 4 caractères.'})
        user = User.objects.create_user(
            username=username,
            password=password,
            first_name=first_name,
            last_name=role,  # store role in last_name
            is_staff=(role == 'admin'),
            is_superuser=False,
        )
        return JsonResponse({'success': True})

@csrf_exempt
@api_role_required('superadmin')
def api_utilisateur_delete(request, user_id):
    if request.method != 'POST':
        return JsonResponse({'success': False}, status=405)
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Utilisateur introuvable.'})
    if user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Impossible de supprimer le super-administrateur.'})
    if user == request.user:
        return JsonResponse({'success': False, 'error': 'Vous ne pouvez pas supprimer votre propre compte.'})
    user.delete()
    return JsonResponse({'success': True})


@api_role_required('superadmin', 'admin', 'caissier')
def api_current_user_info(request):
    """Returns current logged-in user info for the frontend navbar."""
    user = request.user
    role = get_user_role(user)
    display_name = user.get_full_name() or user.username
    # Clean up display name (remove role stored in last_name)
    if user.last_name in ('admin', 'caissier'):
        display_name = user.first_name or user.username
    return JsonResponse({
        'username': user.username,
        'display_name': display_name,
        'role': role,
        'initials': (display_name[:2]).upper() if display_name else 'U',
    })
